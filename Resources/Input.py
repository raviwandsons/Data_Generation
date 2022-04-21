from contextlib import closing
from openpyxl import load_workbook


def get_input_query(file_name) :
    with closing(load_workbook(filename=file_name)) as wb :
        SheetFirst=wb.worksheets[0]
        SheetSecond=wb.worksheets[1]
        SheetFirstMaxRows1=str(SheetFirst.max_row)
        SheetSecondMaxRows=str(SheetSecond.max_row)
        myQuery=[]
        for i in range(1, SheetFirst.max_row) :
            catCode=SheetFirst.cell(row=i + 1, column=1)
            category_code=str(catCode.value)
            catName=SheetFirst.cell(row=i + 1, column=2)
            category_name=str(catName.value)
            for j in range(1, SheetSecond.max_row) :
                cityName=SheetSecond.cell(row=j + 1, column=1)
                printCityName=str(cityName.value)
                stateName=SheetSecond.cell(row=j + 1, column=2)
                printStateName=str(stateName.value)
                query= "All " + category_name + " in " + printCityName + ", " + printStateName + " " + "United States"
                # obj={"query": query, "categorycode": category_code, "filename": category_name+"_"+printCityName+"_"+printStateName+".xlsx"}
                obj={"query": query, "categorycode": category_code, "filename": category_code+"_Category.xlsx"}

                myQuery.append(obj)
        return myQuery
